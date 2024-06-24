import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import csv
import io

colT1, colT2 = st.columns([3, 6])
with colT2:
    st.title(":violet[Luminator]")

# Create a file uploader
with st.popover(
    "Upload Files", help="Upload your streaming report files", use_container_width=True
):
    uploaded_files = st.file_uploader(
        "Upload your streaming report files",
        accept_multiple_files=True,
        type=["csv", "xlsx"],
    )


def excel_to_csv(file):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file, read_only=True)
    # Select the second sheet
    sheet = wb.worksheets[1]

    # Create a CSV writer object
    output = io.StringIO()
    csv_writer = csv.writer(output)

    # Write the rows from the sheet to the CSV
    for row in sheet.iter_rows(values_only=True):
        csv_writer.writerow(row)

    # Seek to the beginning of the StringIO object
    output.seek(0)

    return output


def process_file(file):
    # Determine the file type and read data accordingly
    if file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        try:
            # Convert Excel to CSV using openpyxl and csv
            csv_file = excel_to_csv(file)
            # Read the CSV data
            df = pd.read_csv(csv_file, skiprows=7)  # Assuming header starts at row 7
            df.drop(columns=df.columns[0], inplace=True)
        except Exception as e:
            st.error(f"Error reading Excel file ({file.name}): {e}")
            return None
    else:
        df = pd.read_csv(
            io.BytesIO(file.read()), skiprows=6
        )  # Assuming header starts at row 7

    # Filter out rows where the second column contains "All Airplay Formats Selected"
    df = df[~df.iloc[:, 1].str.contains("All Airplay Formats Selected")]

    if "Range" not in df.columns:
        st.error(
            f"Column 'Range' not found in {file.name}. Please check the file format."
        )
        return None

    # Collect datetime of latest date in rolling date range column
    latest_date = df["Range"][0]

    # Rename columns and clean data
    column_rename = {
        "Range": "Rolling Date Range",
        "TP.7": "This Period Streams",
        "TP.8": "Audio Streams This Period",
        "TP.9": "Video Streams This Period",
        "% Chg.7": "This Period Streams % Change",
        "% Chg.8": "Audio Streams This Period % Change",
        "% Chg.9": "Video Streams This Period % Change",
        "YTD (" + latest_date + ").7": "YTD Streams",
        "YTD (" + latest_date + ").8": "YTD Audio Streams",
        "YTD (" + latest_date + ").9": "YTD Video Streams",
        "ATD (" + latest_date + ").7": "ATD Streams",
        "ATD (" + latest_date + ").8": "ATD Audio Streams",
        "ATD (" + latest_date + ").9": "ATD Video Streams",
        # Add additional column mappings if necessary
    }
    # Drop the 4th column
    df.drop(columns=df.columns[1], inplace=True)
    df.drop(columns=df.columns[3:31], inplace=True)
    df.drop(columns=df.columns[15:], inplace=True)

    df.rename(columns=column_rename, inplace=True)
    df.drop(columns=[col for col in df.columns if "Unnamed" in col], inplace=True)
    df = df.apply(
        lambda x: x.str.strip() if x.dtype == "object" else x
    )  # Strip whitespace from strings
    df = df.apply(
        lambda x: x.str.replace(",", "") if x.dtype == "object" else x
    )  # Remove commas from strings
    return df


# Process each uploaded file
all_data = []
date_ranges = []
for uploaded_file in uploaded_files:
    if uploaded_file is not None:
        df = process_file(uploaded_file)
        if df is not None:
            all_data.append(df)
            # Store the date ranges
            date_ranges.append(
                {
                    "file_name": uploaded_file.name,
                    "start_date": pd.to_datetime(df["Rolling Date Range"].min()),
                    "end_date": pd.to_datetime(df["Rolling Date Range"].max()),
                    "dataframe": df,
                }
            )

# Identify the most recent date range
if date_ranges:
    most_recent_end_date = max([dr["end_date"] for dr in date_ranges])

# Check for non-corresponding date ranges
non_corresponding_files = []
filtered_data = []
for dr in date_ranges:
    if dr["end_date"] < most_recent_end_date:
        non_corresponding_files.append(dr["file_name"])
    else:
        filtered_data.append(dr["dataframe"])

# Display an alert if there are non-corresponding date ranges
if non_corresponding_files:
    st.warning(
        f"The following files have older date ranges and are excluded from the aggregated output: :violet[{', '.join(non_corresponding_files)}]"
    )

    cols1, cols2 = st.columns([2, 3])
    container = cols1.container()
    # Toggle switch to include older date ranges
    include_older_files = container.toggle("INCLUDE OLDER FILES", False)
else:
    include_older_files = False

if include_older_files:
    filtered_data = [dr["dataframe"] for dr in date_ranges]

# Aggregate data
if filtered_data:
    aggregated_data = pd.concat(filtered_data)

    aggregated_data["Rolling Date Range"] = pd.to_datetime(
        aggregated_data["Rolling Date Range"]
    )

    # Initialize an empty DataFrame for the result
    result_df = pd.DataFrame(columns=["Artist"])

    # Group by artist and aggregate data
    for artist, group in aggregated_data.groupby("Artist"):
        # Extract stream, video, audio counts into a list and clean data
        stream_counts = (
            group["This Period Streams"].fillna("0").astype(int).tolist()[::-1]
        )
        audio_counts = (
            group["Audio Streams This Period"].fillna("0").astype(int).tolist()[::-1]
        )
        video_counts = (
            group["Video Streams This Period"].fillna("0").astype(int).tolist()[::-1]
        )

        # Calculate total streams for the specified periods
        tp_total_streams = sum(stream_counts[7:12])
        lp_total_streams = sum(stream_counts[0:7])
        tp_audio_streams = sum(audio_counts[7:12])
        tp_video_streams = sum(video_counts[7:12])
        # Avoid division by zero
        if tp_total_streams == 0:
            audio_percent_change = np.nan
            video_percent_change = np.nan
        else:
            audio_percent_change = (tp_audio_streams / tp_total_streams) * 100
            video_percent_change = (tp_video_streams / tp_total_streams) * 100

        # Avoid division by zero
        if lp_total_streams == 0:
            percent_change = np.nan  # or use 0 or any other placeholder
        else:
            percent_change = (
                (tp_total_streams - lp_total_streams) / lp_total_streams * 100
            )

        # Construct the date range string
        date_range = f"{group['Rolling Date Range'].min().strftime('%A, %d %b %Y')} - {group['Rolling Date Range'].max().strftime('%A, %d %b %Y')}"

        # Create a temporary DataFrame and concat it with result_df
        temp_df = pd.DataFrame(
            {
                "Artist": [artist],
                "Date Range": [date_range],
                "TP Total Streams (Days 8 - 12)": [tp_total_streams],
                "LP Total Streams (Days 1 - 7)": [lp_total_streams],
                "% Change (LP to TP)": [percent_change],
                "TP Audio Streams": [tp_audio_streams],
                "Audio %": [audio_percent_change],
                "TP Video Streams": [tp_video_streams],
                "Video %": [video_percent_change],
                "Stream Counts": [stream_counts],
            }
        )
        result_df = pd.concat([result_df, temp_df], ignore_index=True)

    # Remove date range column
    result_df.drop(columns=["Date Range"], inplace=True)
    period_min = aggregated_data["Rolling Date Range"].min().strftime("%A, %d %b %Y")
    period_max = aggregated_data["Rolling Date Range"].max().strftime("%A, %d %b %Y")

    if len(filtered_data) == 1:
        st.markdown(
            f"***Report for :violet[{len(filtered_data)}] file ({period_min} - {period_max})***"
        )
    elif len(filtered_data) > 1:
        st.markdown(
            f"***Report for :violet[{len(filtered_data)}] files ({period_min} - {period_max})***"
        )

    st.dataframe(
        result_df,
        use_container_width=True,
        hide_index=True,
        column_config={"Stream Counts": st.column_config.AreaChartColumn()},
    )

    # Remove the last column of the DataFrame
    result_df = result_df.iloc[:, :-1]

    # Convert the DataFrame to an Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False)
    # Load the workbook and adjust the column widths
    output.seek(0)
    book = openpyxl.load_workbook(output)
    sheet = book.active
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook to a new BytesIO object
    output = io.BytesIO()
    book.save(output)
    output.seek(0)

    # Create the download button
    if non_corresponding_files:
        cols2.download_button(
            label="Download data as Excel",
            data=output,
            file_name="data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.download_button(
            label="Download data as Excel",
            data=output,
            file_name="data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
